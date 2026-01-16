"""
OCRimageflow - FastAPI Backend
Sistema de procesamiento de imágenes con Google Vision, Gemini AI y normalización inteligente
"""

from fastapi import FastAPI, HTTPException, Depends, status, UploadFile, File, Form
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, EmailStr
from typing import Optional, List, Dict
from datetime import datetime, timedelta
import jwt
import bcrypt
import os
from dotenv import load_dotenv
import psycopg2
from psycopg2.extras import RealDictCursor
import boto3
from botocore.exceptions import ClientError
import uuid
import base64
import json
import mimetypes
import requests
import re
from google.cloud import vision
import io
from PIL import Image
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
import tempfile

load_dotenv()

app = FastAPI(title="OCRimageflow API", version="1.0.0")

# CORS Configuration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Security
security = HTTPBearer()
SECRET_KEY = os.getenv("JWT_SECRET_KEY", "your-secret-key-change-in-production")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 7  # 7 días

# API Keys
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
GOOGLE_CREDENTIALS_JSON = os.getenv("GOOGLE_CREDENTIALS_JSON")  # Base64 encoded

# Database Configuration
DATABASE_URL = os.getenv("DATABASE_URL")

# AWS S3 Configuration
AWS_ACCESS_KEY_ID = os.getenv("AWS_ACCESS_KEY_ID")
AWS_SECRET_ACCESS_KEY = os.getenv("AWS_SECRET_ACCESS_KEY")
AWS_BUCKET_NAME = os.getenv("AWS_BUCKET_NAME")
AWS_REGION = os.getenv("AWS_REGION", "us-east-1")

# Initialize S3 client
s3_client = boto3.client(
    's3',
    aws_access_key_id=AWS_ACCESS_KEY_ID,
    aws_secret_access_key=AWS_SECRET_ACCESS_KEY,
    region_name=AWS_REGION
)

# Initialize Google Vision (decode credentials from env)
if GOOGLE_CREDENTIALS_JSON:
    creds_dict = json.loads(base64.b64decode(GOOGLE_CREDENTIALS_JSON))
    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as f:
        json.dump(creds_dict, f)
        os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = f.name

vision_client = vision.ImageAnnotatorClient()

# ===========================================
# DICCIONARIOS DE NORMALIZACIÓN
# ===========================================

FIELD_NORMALIZATION = {
    "pre$io": "precio_unitario", "precio": "precio_unitario", "price": "precio_unitario",
    "costo": "precio_unitario", "cost": "precio_unitario", "valor": "precio_unitario",
    "pr3cio": "precio_unitario", "preclo": "precio_unitario",
    "talla": "talla", "size": "talla", "medida": "talla", "s1ze": "talla",
    "tamano": "talla", "tamaño": "talla",
    "peso": "peso", "weight": "peso", "piso": "peso", "pezo": "peso",
    "color": "color", "colour": "color", "c0lor": "color",
    "composicion": "composicion_textil", "composición": "composicion_textil",
    "composition": "composicion_textil", "material": "composicion_textil",
    "fabric": "composicion_textil", "tela": "composicion_textil",
    "alto": "alto", "altura": "alto", "height": "alto", "h": "alto",
    "ancho": "ancho", "width": "ancho", "w": "ancho", "widht": "ancho",
    "largo": "largo", "length": "largo", "profundidad": "largo", "l": "largo",
    "pecho": "medida_pecho", "chest": "medida_pecho", "bust": "medida_pecho",
    "brazo": "medida_brazo", "manga": "medida_brazo", "sleeve": "medida_brazo",
    "cuello": "medida_cuello", "neck": "medida_cuello", "collar": "medida_cuello",
    "espalda": "medida_espalda", "back": "medida_espalda", "shoulder": "medida_espalda",
    "qty": "qty_por_caja", "quantity": "qty_por_caja", "cantidad": "qty_por_caja",
    "pcs": "qty_por_caja", "piezas": "qty_por_caja", "unidades": "qty_por_caja",
    "cbm": "cbm_por_caja", "volumen": "cbm_por_caja", "volume": "cbm_por_caja", "m3": "cbm_por_caja",
    "sku": "sku", "codigo": "sku", "code": "sku", "ref": "sku",
    "genero": "genero", "género": "genero", "gender": "genero",
    "sexo": "genero", "nino": "genero", "niño": "genero", "nina": "genero", "niña": "genero",
    "edad": "edad_rango", "age": "edad_rango", "meses": "edad_rango", "months": "edad_rango",
    "suela": "composicion_suela", "sole": "composicion_suela",
    "interior": "composicion_interior", "lining": "composicion_interior",
    "exterior": "composicion_exterior", "upper": "composicion_exterior",
    "cordones": "cordones", "laces": "cordones",
    "tipo": "tipo_tela", "tipo_tela": "tipo_tela",
    "yardas": "yardas_rollo", "yards": "yardas_rollo",
    "metros": "metros_rollo", "meters": "metros_rollo",
    "gama": "gama", "calidad": "gama", "quality": "gama",
    "marca": "marca", "brand": "marca",
    "modelo": "modelo", "model": "modelo"
}

UNIT_CORRECTIONS = {
    "KF": "kg", "kf": "kg", "KG": "kg", "Kg": "kg",
    "LB": "lb", "lbs": "lb", "libras": "lb",
    "CM": "cm", "M": "m", "IN": "in",
    "M3": "m³", "CBM": "m³"
}

INDUSTRY_KEYWORDS = {
    "fashion": ["camisa", "pantalon", "ropa", "talla", "composicion", "shirt", "clothing", "fabric"],
    "furniture": ["mueble", "silla", "mesa", "furniture", "chair", "alto", "ancho"],
    "footwear": ["zapato", "bota", "shoe", "suela", "sole"],
    "baby": ["bebe", "baby", "infantil", "meses"],
    "textile": ["tela", "textil", "rollo", "fabric", "yardas"]
}

INDUSTRY_COLUMN_ORDER = {
    "fashion": ["sku", "composicion_textil", "talla", "color", "peso", "precio_unitario"],
    "furniture": ["sku", "material", "alto", "largo", "ancho", "precio_unitario"],
    "footwear": ["sku", "talla", "composicion_suela", "color", "precio_unitario"],
    "baby": ["sku", "edad_rango", "genero", "color", "precio_unitario"],
    "textile": ["sku", "tipo_tela", "metros_rollo", "precio_unitario"]
}

# Tier configurations
TIER_CONFIGS = {
    "free": {
        "max_images": 10,
        "max_images_per_batch": 5,
        "ocr_engine": "google_vision",
        "retention_days": 3,
        "max_suppliers": 1
    },
    "starter": {
        "max_images": 200,
        "max_images_per_batch": 50,
        "ocr_engine": "google_vision",
        "retention_days": 30,
        "max_suppliers": 3
    },
    "basic": {
        "max_images": 500,
        "max_images_per_batch": 100,
        "ocr_engine": "google_vision",
        "retention_days": 30,
        "max_suppliers": 3
    },
    "pro": {
        "max_images": 2000,
        "max_images_per_batch": 200,
        "ocr_engine": "gemini",
        "retention_days": 90,
        "max_suppliers": 5
    },
    "enterprise": {
        "max_images": 10000,
        "max_images_per_batch": 500,
        "ocr_engine": "gemini",
        "retention_days": 90,
        "max_suppliers": 999
    }
}

# ===========================================
# DATABASE CONNECTION
# ===========================================

def get_db_connection():
    conn = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
    return conn

# ===========================================
# PYDANTIC MODELS
# ===========================================

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    tier: str = "free"

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class ProcessResponse(BaseModel):
    status: str
    images_processed: int
    industry_detected: str
    excel_url: Optional[str] = None
    normalized_data: List[Dict]

# ===========================================
# HELPER FUNCTIONS
# ===========================================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
        return int(user_id)
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid token")

def upload_to_s3(file_content: bytes, filename: str, user_id: int, content_type: str) -> str:
    """Upload file to S3 and return the URL"""
    try:
        unique_filename = f"user_{user_id}/{uuid.uuid4()}_{filename}"
        
        s3_client.put_object(
            Bucket=AWS_BUCKET_NAME,
            Key=unique_filename,
            Body=file_content,
            ContentType=content_type
        )
        
        return f"https://{AWS_BUCKET_NAME}.s3.{AWS_REGION}.amazonaws.com/{unique_filename}"
    except ClientError as e:
        raise HTTPException(status_code=500, detail=f"S3 upload failed: {str(e)}")

def log_usage(user_id: int, action: str, details: Optional[dict] = None):
    """Log user actions to database"""
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "INSERT INTO usage_logs (user_id, action, details, images_processed, cost) VALUES (%s, %s, %s, %s, %s)",
            (user_id, action, json.dumps(details) if details else None, 
             details.get('images_processed', 0) if details else 0,
             details.get('cost', 0.0) if details else 0.0)
        )
        conn.commit()
    finally:
        cursor.close()
        conn.close()

# ===========================================
# DATA NORMALIZER
# ===========================================

class DataNormalizer:
    def __init__(self):
        self.field_map = FIELD_NORMALIZATION
        self.unit_map = UNIT_CORRECTIONS
    
    def detect_industry(self, raw_data):
        text = " ".join([str(v).lower() for v in raw_data.values()])
        scores = {industry: sum(1 for kw in keywords if kw in text) 
                  for industry, keywords in INDUSTRY_KEYWORDS.items()}
        return max(scores, key=scores.get) if max(scores.values()) > 0 else "general"
    
    def normalize_field_name(self, raw_name):
        clean = raw_name.lower().strip().replace("$", "").replace(":", "").replace("_", " ").strip()
        for key, normalized in self.field_map.items():
            if key in clean or clean in key:
                return normalized
        return clean.replace(" ", "_")
    
    def normalize_value(self, field_name, value, industry="general"):
        if not value:
            return ""
        value_str = str(value).strip()
        
        if "precio" in field_name or "price" in field_name:
            numbers = re.findall(r'\d+\.?\d*', value_str)
            return f"${float(numbers[0]):.2f}" if numbers else value_str
        
        elif "peso" in field_name or "weight" in field_name:
            for wrong, correct in self.unit_map.items():
                if wrong.lower() in value_str.lower():
                    value_str = value_str.lower().replace(wrong.lower(), correct)
            match = re.search(r'(\d+\.?\d*)\s*([a-zA-Z]+)', value_str)
            return f"{match.group(1)} {self.unit_map.get(match.group(2).upper(), match.group(2).lower())}" if match else value_str
        
        elif "talla" in field_name or "size" in field_name:
            size_map = {"S": "S", "M": "M", "L": "L", "XL": "XL", "XXL": "XXL"}
            return size_map.get(value_str.upper(), value_str.upper())
        
        elif "color" in field_name:
            return value_str.capitalize()
        
        return value_str.capitalize() if value_str else value_str
    
    def normalize_data(self, raw_data, industry=None):
        if not industry:
            industry = self.detect_industry(raw_data)
        
        normalized = {}
        for raw_key, raw_value in raw_data.items():
            if raw_key.startswith("_"):
                continue
            field_name = self.normalize_field_name(raw_key)
            normalized[field_name] = self.normalize_value(field_name, raw_value, industry)
        
        return normalized, industry

# ===========================================
# OCR ENGINES
# ===========================================

def google_vision_ocr(image_bytes: bytes) -> dict:
    """Extract text using Google Vision API"""
    try:
        image = vision.Image(content=image_bytes)
        response = vision_client.text_detection(image=image)
        
        if response.error.message:
            raise Exception(response.error.message)
        
        texts = response.text_annotations
        if not texts:
            return {"text": "", "confidence": 0, "structured_data": {}}
        
        full_text = texts[0].description
        confidence = texts[0].score if hasattr(texts[0], 'score') else 0.9
        
        return {
            "text": full_text,
            "confidence": confidence,
            "structured_data": parse_text_to_dict(full_text),
            "engine": "google_vision"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Google Vision OCR failed: {str(e)}")

def gemini_ocr(image_bytes: bytes, mime_type: str) -> dict:
    """Extract text using Gemini AI"""
    try:
        b64 = base64.b64encode(image_bytes).decode()
        url = "https://generativelanguage.googleapis.com/v1/models/gemini-2.0-flash-exp:generateContent"
        payload = {
            "contents": [{
                "parts": [
                    {"text": "Extrae TODOS los datos de esta imagen en formato JSON. Usa claves descriptivas en español. Ejemplo: {\"precio\": \"$10.50\", \"talla\": \"M\", \"color\": \"Azul\"}"},
                    {"inline_data": {"mime_type": mime_type, "data": b64}}
                ]
            }]
        }
        
        r = requests.post(f"{url}?key={GEMINI_API_KEY}", json=payload, timeout=30)
        r.raise_for_status()
        
        content = r.json()["candidates"][0]["content"]["parts"][0]["text"]
        content = content.replace("```json", "").replace("```", "").strip()
        data = json.loads(content)
        
        return {
            "text": json.dumps(data, ensure_ascii=False),
            "confidence": 0.95,
            "structured_data": data,
            "engine": "gemini"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Gemini OCR failed: {str(e)}")

def parse_text_to_dict(text: str) -> dict:
    """Parse plain text to dictionary"""
    fields = {}
    for line in text.split('\n'):
        if ':' in line:
            parts = line.split(':', 1)
            if len(parts) == 2 and parts[0].strip() and parts[1].strip():
                fields[parts[0].strip()] = parts[1].strip()
    return fields

# ===========================================
# EXCEL GENERATION
# ===========================================

def generate_excel(data_list: List[dict], image_urls: List[str], industry: str, user_id: int) -> bytes:
    """Generate Excel file with images and normalized data"""
    column_order = INDUSTRY_COLUMN_ORDER.get(industry, [])
    all_fields = set()
    for item in data_list:
        all_fields.update(k for k in item.keys() if k != '_metadata')
    
    ordered_fields = [c for c in column_order if c in all_fields]
    ordered_fields.extend(sorted(all_fields - set(ordered_fields)))
    
    rows = [{"Imagen": "", **{f: item.get(f, "") for f in ordered_fields}} for item in data_list]
    df = pd.DataFrame(rows)
    
    output = BytesIO()
    df.to_excel(output, sheet_name="Datos", index=False, engine='openpyxl')
    output.seek(0)
    
    wb = load_workbook(output)
    ws = wb["Datos"]
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws.column_dimensions["A"].width = 25
    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Add images
    for idx, img_url in enumerate(image_urls):
        try:
            # Download image from S3
            response = requests.get(img_url, timeout=10)
            if response.status_code == 200:
                img = Image.open(BytesIO(response.content))
                img.thumbnail((150, 150), Image.Resampling.LANCZOS)
                img_io = BytesIO()
                img.save(img_io, format='PNG')
                img_io.seek(0)
                xl_img = XLImage(img_io)
                ws.row_dimensions[idx + 2].height = 120
                xl_img.anchor = f"A{idx + 2}"
                ws.add_image(xl_img)
        except:
            pass
    
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output.getvalue()

# ===========================================
# ROUTES
# ===========================================

@app.get("/")
def read_root():
    return {
        "message": "OCRimageflow API",
        "version": "1.0.0",
        "status": "online",
        "features": ["Google Vision OCR", "Gemini AI", "Smart Normalization", "Industry Detection"]
    }

@app.get("/health")
def health_check():
    """Check if the API and database are working"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT 1")
        cursor.close()
        conn.close()
        return {"status": "healthy", "database": "connected", "ocr": "ready"}
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Service unavailable: {str(e)}")

@app.post("/auth/register")
def register(user: UserCreate):
    """Register a new user"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Validate tier
        if user.tier not in TIER_CONFIGS:
            raise HTTPException(status_code=400, detail="Invalid tier")
        
        # Check if user exists
        cursor.execute("SELECT id FROM users WHERE email = %s", (user.email,))
        if cursor.fetchone():
            raise HTTPException(status_code=400, detail="Email already registered")
        
        # Create user
        hashed_password = hash_password(user.password)
        cursor.execute(
            "INSERT INTO users (email, password_hash, name, tier) VALUES (%s, %s, %s, %s) RETURNING id",
            (user.email, hashed_password, user.name, user.tier)
        )
        user_id = cursor.fetchone()['id']
        conn.commit()
        
        # Create token
        token = create_access_token({"sub": str(user_id)})
        
        log_usage(user_id, "user_registered", {"tier": user.tier})
        
        return {
            "access_token": token,
            "token_type": "bearer",
            "user": {
                "id": user_id,
                "email": user.email,
                "name": user.name,
                "tier": user.tier
            }
        }
    finally:
        cursor.close()
        conn.close()

@app.post("/auth/login")
def login(credentials: UserLogin):
    """Login user"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute(
            "SELECT id, email, name, password_hash, tier FROM users WHERE email = %s",
            (credentials.email,)
        )
        user = cursor.fetchone()
        
        if not user or not verify_password(credentials.password, user['password_hash']):
            raise HTTPException(status_code=401, detail="Invalid credentials")
        
        token = create_access_token({"sub": str(user['id'])})
        
        log_usage(user['id'], "user_login")
        
        return {
            "access_token": token,
            "token_type": "bearer",
            "user": {
                "id": user['id'],
                "email": user['email'],
                "name": user['name'],
                "tier": user['tier']
            }
        }
    finally:
        cursor.close()
        conn.close()

@app.post("/process/batch")
async def process_batch(
    files: List[UploadFile] = File(...),
    user_id: int = Depends(get_current_user)
):
    """Process multiple images with OCR and normalization"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Get user tier
        cursor.execute("SELECT tier, images_processed_this_month FROM users WHERE id = %s", (user_id,))
        user = cursor.fetchone()
        
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        tier = user['tier']
        config = TIER_CONFIGS[tier]
        images_this_month = user['images_processed_this_month'] or 0
        
        # Check limits
        if len(files) > config['max_images_per_batch']:
            raise HTTPException(
                status_code=400, 
                detail=f"Batch too large. Max {config['max_images_per_batch']} images per batch for {tier} tier"
            )
        
        if images_this_month + len(files) > config['max_images']:
            raise HTTPException(
                status_code=403,
                detail=f"Monthly limit exceeded. {images_this_month}/{config['max_images']} images used"
            )
        
        # Process images
        normalizer = DataNormalizer()
        extracted_data = []
        image_urls = []
        industries = []
        
        for file in files:
            # Read image
            image_bytes = await file.read()
            
            # Upload to S3
            image_url = upload_to_s3(
                image_bytes,
                file.filename,
                user_id,
                file.content_type or 'image/jpeg'
            )
            image_urls.append(image_url)
            
            # OCR
            if config['ocr_engine'] == 'google_vision':
                ocr_result = google_vision_ocr(image_bytes)
            else:
                ocr_result = gemini_ocr(image_bytes, file.content_type or 'image/jpeg')
            
            # Normalize
            raw_data = ocr_result['structured_data']
            normalized, industry = normalizer.normalize_data(raw_data)
            industries.append(industry)
            
            normalized['_metadata'] = {
                "image_url": image_url,
                "industry": industry,
                "ocr_engine": ocr_result['engine']
            }
            extracted_data.append(normalized)
        
        # Detect main industry
        main_industry = max(set(industries), key=industries.count) if industries else "general"
        
        # Generate Excel
        excel_bytes = generate_excel(extracted_data, image_urls, main_industry, user_id)
        
        # Upload Excel to S3
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"batch_{tier}_{main_industry}_{timestamp}.xlsx"
        excel_url = upload_to_s3(excel_bytes, excel_filename, user_id, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # Update user stats
        cursor.execute(
            "UPDATE users SET images_processed_this_month = images_processed_this_month + %s WHERE id = %s",
            (len(files), user_id)
        )
        conn.commit()
        
        # Log usage
        log_usage(user_id, "batch_processed", {
            "images_processed": len(files),
            "industry": main_industry,
            "tier": tier,
            "cost": 0.0
        })
        
        return {
            "status": "success",
            "images_processed": len(files),
            "industry_detected": main_industry,
            "excel_url": excel_url,
            "normalized_data": extracted_data,
            "remaining_images": config['max_images'] - (images_this_month + len(files))
        }
        
    finally:
        cursor.close()
        conn.close()

@app.get("/usage/stats")
def get_usage_stats(user_id: int = Depends(get_current_user)):
    """Get usage statistics for the current user"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute(
            """SELECT tier, images_processed_this_month, created_at 
               FROM users WHERE id = %s""",
            (user_id,)
        )
        user = cursor.fetchone()
        
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        config = TIER_CONFIGS[user['tier']]
        
        cursor.execute(
            """SELECT COUNT(*) as total_batches, SUM(images_processed) as total_images
               FROM usage_logs WHERE user_id = %s AND action = 'batch_processed'""",
            (user_id,)
        )
        stats = cursor.fetchone()
        
        return {
            "tier": user['tier'],
            "images_this_month": user['images_processed_this_month'],
            "max_images_per_month": config['max_images'],
            "max_images_per_batch": config['max_images_per_batch'],
            "remaining_images": config['max_images'] - (user['images_processed_this_month'] or 0),
            "total_batches_processed": stats['total_batches'] or 0,
            "total_images_processed": stats['total_images'] or 0,
            "member_since": user['created_at']
        }
    finally:
        cursor.close()
        conn.close()

@app.get("/usage/logs")
def get_usage_logs(user_id: int = Depends(get_current_user), limit: int = 50):
    """Get usage logs for the current user"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute(
            """SELECT id, action, details, images_processed, cost, created_at 
               FROM usage_logs 
               WHERE user_id = %s 
               ORDER BY created_at DESC 
               LIMIT %s""",
            (user_id, limit)
        )
        logs = cursor.fetchall()
        
        return {"logs": logs}
    finally:
        cursor.close()
        conn.close()

@app.get("/tiers")
def get_tiers():
    """Get available tiers and their limits"""
    return {
        "tiers": {
            tier: {
                "max_images_per_month": config['max_images'],
                "max_images_per_batch": config['max_images_per_batch'],
                "ocr_engine": config['ocr_engine']
            }
            for tier, config in TIER_CONFIGS.items()
        }
    }

# ===========================================
# SUPPLIERS ENDPOINTS
# ===========================================

def create_thumbnail(image_bytes: bytes, max_size: tuple = (300, 300)) -> bytes:
    """Create thumbnail from image bytes"""
    try:
        img = Image.open(io.BytesIO(image_bytes))
        img.thumbnail(max_size, Image.Resampling.LANCZOS)
        
        output = io.BytesIO()
        img.save(output, format='JPEG', quality=85)
        output.seek(0)
        return output.read()
    except Exception as e:
        print(f"Error creating thumbnail: {e}")
        return None

@app.post("/suppliers")
async def create_supplier(
    name: str = Form(...),
    description: str = Form(None),
    current_user: dict = Depends(get_current_user)
):
    """Create a new supplier/provider"""
    user_id = current_user['id']
    tier = current_user['tier']
    tier_config = TIER_CONFIGS.get(tier, TIER_CONFIGS['free'])
    max_suppliers = tier_config.get('max_suppliers', 1)
    
    conn = psycopg2.connect(DATABASE_URL)
    cursor = conn.cursor()
    
    try:
        cursor.execute(
            "SELECT COUNT(*) FROM suppliers WHERE user_id = %s AND is_active = true",
            (user_id,)
        )
        current_count = cursor.fetchone()[0]
        
        if current_count >= max_suppliers:
            raise HTTPException(
                status_code=403,
                detail=f"Your {tier} plan allows maximum {max_suppliers} suppliers. Upgrade to add more."
            )
        
        cursor.execute(
            "SELECT id FROM suppliers WHERE user_id = %s AND name = %s",
            (user_id, name)
        )
        if cursor.fetchone():
            raise HTTPException(status_code=400, detail=f"Supplier '{name}' already exists")
        
        cursor.execute(
            """
            INSERT INTO suppliers (user_id, name, description)
            VALUES (%s, %s, %s)
            RETURNING id, name, description, total_images_processed, is_active, created_at
            """,
            (user_id, name, description)
        )
        row = cursor.fetchone()
        conn.commit()
        
        return {
            "id": row[0],
            "name": row[1],
            "description": row[2],
            "total_images_processed": row[3],
            "is_active": row[4],
            "created_at": row[5]
        }
    
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()

@app.get("/suppliers")
async def list_suppliers(current_user: dict = Depends(get_current_user)):
    """List all active suppliers for current user"""
    user_id = current_user['id']
    
    conn = psycopg2.connect(DATABASE_URL)
    cursor = conn.cursor()
    
    try:
        cursor.execute(
            """
            SELECT id, name, description, total_images_processed, is_active, created_at
            FROM suppliers
            WHERE user_id = %s AND is_active = true
            ORDER BY created_at DESC
            """,
            (user_id,)
        )
        rows = cursor.fetchall()
        
        return {
            "suppliers": [
                {
                    "id": row[0],
                    "name": row[1],
                    "description": row[2],
                    "total_images_processed": row[3],
                    "is_active": row[4],
                    "created_at": row[5]
                }
                for row in rows
            ]
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()

@app.get("/suppliers/{supplier_id}/stats")
async def get_supplier_stats(supplier_id: int, current_user: dict = Depends(get_current_user)):
    """Get statistics for a supplier"""
    user_id = current_user['id']
    
    conn = psycopg2.connect(DATABASE_URL)
    cursor = conn.cursor()
    
    try:
        cursor.execute(
            """
            SELECT * FROM supplier_stats
            WHERE supplier_id = %s AND user_id = %s
            """,
            (supplier_id, user_id)
        )
        row = cursor.fetchone()
        
        if not row:
            raise HTTPException(status_code=404, detail="Supplier not found")
        
        return {
            "supplier_id": row[0],
            "supplier_name": row[2],
            "total_batches": row[4] if len(row) > 4 else 0,
            "total_images": row[5] if len(row) > 5 else 0,
            "images_processed": row[6] if len(row) > 6 else 0,
            "last_batch_date": row[7] if len(row) > 7 and row[7] else None
        }
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
